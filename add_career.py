import json,re,math,statistics,hashlib
from pathlib import Path
from collections import Counter,defaultdict
from openpyxl import load_workbook
BASE=Path('/home/user/drive_data')
OUT=Path('/home/user/analysis_output/analysis_data.json')
D=json.load(open(OUT,encoding='utf-8'))

def norm(x):
    if x is None: return ''
    return re.sub(r'\s+','',str(x).strip().upper())
def norm_text(x): return '' if x is None else str(x).strip()
def anon(*parts): return hashlib.sha1('|'.join(map(str,parts)).encode()).hexdigest()[:6].upper()
def num(x):
    if x is None: return None
    if isinstance(x,(int,float)) and not isinstance(x,bool): return float(x) if not math.isnan(x) else None
    if isinstance(x,str):
        s=x.strip().upper()
        if s=='U': return 0.0
        if s in ('A','ATTAINED'): return 'A'
        m=re.match(r'^(\d+(?:\.\d+)?)',s)
        return float(m.group(1)) if m else None
    return None

def stats(vals):
    vals=[v for v in vals if isinstance(v,(int,float))]
    if not vals: return {'n':0,'mean':None,'median':None}
    return {'n':len(vals),'mean':round(sum(vals)/len(vals),2),'median':round(statistics.median(vals),2)}

def pct(a,b): return round(100*a/b,1) if b else None

# Parse career files
career=[]
for f in sorted((BASE/'Career Deveolpment').glob('*.xlsx')):
    m=re.search(r'(20\d{2})-(20\d{2})',f.name)
    grad_year=int(m.group(2)) if m else None
    wb=load_workbook(f,data_only=True,read_only=False)
    ws=wb['表單回應 1']
    headers=[c.value for c in ws[1]]
    for rid,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        if not row or not row[2]: continue
        status=norm_text(row[4] if len(row)>4 else '')
        location=norm_text(row[5] if len(row)>5 else '')
        course_cat=norm_text(row[6] if len(row)>6 else '')
        pathway=norm_text(row[7] if len(row)>7 else '')
        inst=norm_text(row[8] if len(row)>8 else '')
        programme=norm_text(row[9] if len(row)>9 else '')
        repeat_form=norm_text(row[10] if len(row)>10 else '')
        industry=norm_text(row[25] if len(row)>25 else '')
        # If first set blank, scan alternate repeated groups for institution/course/category
        cats=[]; insts=[]; progs=[]
        for ci in [6,13,16,19,23]:
            if ci<len(row) and row[ci]: cats.append(norm_text(row[ci]))
        for ii in [8,12,15,18,22]:
            if ii<len(row) and row[ii]: insts.append(norm_text(row[ii]))
        for pi in [9,14,17,20,24]:
            if pi<len(row) and row[pi]: progs.append(norm_text(row[pi]))
        if not course_cat and cats: course_cat=cats[0]
        if not inst and insts: inst=insts[0]
        if not programme and progs: programme=progs[0]
        if not location and len(row)>21 and row[21]: location=norm_text(row[21])
        # standardize status
        status_std='未知/未分類'
        if '升學' in status: status_std='升學'
        elif '就業' in status: status_std='就業'
        elif '重讀' in status or repeat_form: status_std='重讀'
        elif '其他' in status: status_std='其他'
        # standardize level/category
        cat=course_cat
        level='其他/未分類'
        if any(k in cat for k in ['高級文憑','Higher Diploma','HIGHER DIPLOMA','HD']) or any(k in programme for k in ['高級文憑','Higher Diploma','HIGHER DIPLOMA']): level='高級文憑'
        elif any(k in cat for k in ['副學士','Associate','ASSOCIATE']) or any(k in programme for k in ['副學士','Associate','ASSOCIATE']): level='副學士'
        elif any(k in cat for k in ['學士','大學','JUPAS','SSSDP']) or any(k in pathway for k in ['JUPAS','聯招','SSSDP']): level='學士/本科'
        elif any(k in cat for k in ['基礎文憑','職專文憑','文憑']) and '應用教育' not in cat and '高級文憑' not in cat: level='文憑/基礎/職專'
        elif any(k in cat for k in ['應用教育','DAE','毅進']): level='應用教育/毅進'
        elif status_std=='重讀': level='重讀'
        elif status_std=='就業': level='就業'
        elif location and location not in ['本地','香港']: level='海外/其他地區升學'
        # institution type
        inst_type='其他/未分類'
        if any(k in pathway+inst for k in ['大學聯招','JUPAS','香港大學','中文大學','科技大學','理工大學','城市大學','浸會大學','嶺南大學','教育大學','都會大學']) and level=='學士/本科': inst_type='本地學位/JUPAS或SSSDP'
        elif 'VTC' in pathway or any(k in inst for k in ['IVE','香港知專設計學院','中華廚藝學院','青年學院','職業訓練局']): inst_type='VTC/職專'
        elif any(k in pathway+inst for k in ['自資','社區學院','專上學院','持續教育','專業進修','明愛','恒生','珠海','東華學院','伍倫貢','香港科技專上']): inst_type='自資專上/社區學院'
        elif any(k in pathway+cat for k in ['應用教育','DAE','毅進']): inst_type='DAE/毅進'
        elif status_std=='就業': inst_type='就業'
        elif status_std=='重讀': inst_type='重讀'
        # field from programme keywords
        txt=(programme+' '+inst+' '+industry).lower()
        field='其他/未分類'
        field_map=[('護理/醫療/健康',['護理','nursing','健康','health','醫療','救護','幼兒','社會工作','social work','心理']),('工程/建築/設計',['工程','建築','土木','電機','機械','測量','設計','design','室內']),('商業/會計/管理',['會計','商業','管理','business','財務','marketing','旅遊','酒店','康樂','金融']),('資訊科技/數碼媒體',['電腦','資訊','ict','數碼','媒體','電影','音樂製作','遊戲','ai','data']),('教育/社會服務',['教育','幼兒','社會工作','社工','community','social']),('體育/紀律/服務',['警務','懲教','消防','紀律','運動','教練','航空','服務']),('語文/人文/藝術',['中文','英文','翻譯','歷史','藝術','音樂','art','creative'])]
        for name,keys in field_map:
            if any(k.lower() in txt for k in keys): field=name; break
        career.append({'grad_year':grad_year,'cohort_label':f'{grad_year-1}-{str(grad_year)[-2:]}' if grad_year else '', 'row':rid, 'class':norm_text(row[1]), 'cname':norm(row[2]), 'anon':'CD-'+anon(grad_year,row[1],row[2]), 'status_raw':status, 'status':status_std, 'location':location or '未填', 'course_category_raw':course_cat, 'level':level, 'pathway':pathway or '未填', 'institution':inst or '未填', 'institution_type':inst_type, 'programme':programme or '未填', 'industry':industry or '未填', 'field':field})

# Parse DSE student records 2022-2025 with names
DSE_FILES=[]
for y in [2022,2023,2024,2025]:
    p=BASE/f'HKDSE/各班成績/各班成績{y}.xlsx'
    if not p.exists(): p=BASE/f'HKDSE/Reference/{y}/各班成績{y}.xlsx'
    DSE_FILES.append((y,p))

def subject_columns(ws):
    header_row=None
    for r in range(1,min(6,ws.max_row)+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if '班別' in vals: header_row=r; break
    if not header_row: return None,[]
    hdr=[ws.cell(header_row,c).value for c in range(1,ws.max_column+1)]
    grp=[ws.cell(header_row-1,c).value if header_row>1 else None for c in range(1,ws.max_column+1)]
    cols=[]
    for i,h in enumerate(hdr):
        hstr=str(h).strip() if h is not None else ''; gstr=str(grp[i]).strip() if grp[i] is not None else ''
        subj=None
        if gstr.upper()=='CHINESE' and hstr in ['整體','Overall']: subj='中文'
        elif gstr.upper()=='ENGLISH' and hstr in ['Overall','整體','TOTAL']: subj='英文'
        elif hstr in ['數學','公民','M1','M2','會計','商管','生物','化學','物理','中史','歷史','地理','旅款','經濟','ICT','BAFS','企會財','企管']: subj=hstr
        if subj: cols.append((i,subj))
    return header_row,cols

dse=[]
for year,p in DSE_FILES:
    if not p.exists(): continue
    wb=load_workbook(p,data_only=True,read_only=False); ws=wb[wb.sheetnames[0]]
    hr,cols=subject_columns(ws)
    if not hr: continue
    for row in ws.iter_rows(min_row=hr+1,values_only=True):
        if not row or not row[0] or not row[3]: continue
        scores={}
        for idx,subj in cols:
            v=num(row[idx] if idx<len(row) else None)
            if isinstance(v,(int,float)): scores[subj]=v
            elif v=='A': scores[subj]='A'
        numeric=[v for v in scores.values() if isinstance(v,(int,float))]
        best5=sum(sorted(numeric,reverse=True)[:5]) if len(numeric)>=5 else None
        chi=scores.get('中文'); eng=scores.get('英文'); math_score=scores.get('數學')
        other2=sum(1 for k,v in scores.items() if k not in ['中文','英文','數學'] and isinstance(v,(int,float)) and v>=2)
        meet3322=isinstance(chi,(int,float)) and chi>=3 and isinstance(eng,(int,float)) and eng>=3 and isinstance(math_score,(int,float)) and math_score>=2 and other2>=2
        dse.append({'year':year,'class':str(row[0]),'no':row[1],'ename':norm(row[2]),'cname':norm(row[3]),'sex':row[4] if len(row)>4 else '', 'scores':scores,'best5':best5,'meet3322':meet3322,'core_ok': isinstance(chi,(int,float)) and chi>=3 and isinstance(eng,(int,float)) and eng>=3 and isinstance(math_score,(int,float)) and math_score>=2})

idx=defaultdict(list)
for s in dse: idx[(s['year'],s['cname'])].append(s)
linked=[]; unlinked=[]
for c in career:
    cand=idx.get((c['grad_year'],c['cname']),[])
    if len(cand)==1:
        s=cand[0]
        linked.append({**c,'dse_best5':s['best5'],'meet3322':s['meet3322'],'core_ok':s['core_ok'],'dse_class_no':f"{s['class']}#{int(s['no']) if isinstance(s['no'],(int,float)) else s['no']}", 'dse_chi':s['scores'].get('中文'), 'dse_eng':s['scores'].get('英文'), 'dse_math':s['scores'].get('數學')})
    else:
        unlinked.append(c)

# Summaries
career_summary=[]
for y in sorted(set(c['grad_year'] for c in career)):
    arr=[c for c in career if c['grad_year']==y]
    total=len(arr)
    for dim in ['status','level','institution_type','field','location']:
        for k,v in Counter(c[dim] for c in arr).most_common():
            career_summary.append({'grad_year':y,'dimension':dim,'category':k,'count':v,'pct':pct(v,total)})
# linked summaries by DSE bands
bands=[('0-9',0,9),('10-14',10,14),('15-19',15,19),('20+',20,999)]
outcome_by_dse=[]
for name,lo,hi in bands:
    arr=[x for x in linked if x['dse_best5'] is not None and lo<=x['dse_best5']<=hi]
    for dim in ['level','institution_type','field']:
        total=len(arr)
        for k,v in Counter(x[dim] for x in arr).most_common():
            outcome_by_dse.append({'best5_band':name,'dimension':dim,'category':k,'count':v,'pct':pct(v,total),'n_band':total})
for flag,label in [(True,'3322達標'),(False,'3322未達標')]:
    arr=[x for x in linked if x.get('meet3322')==flag]
    total=len(arr)
    for k,v in Counter(x['level'] for x in arr).most_common():
        outcome_by_dse.append({'best5_band':label,'dimension':'level','category':k,'count':v,'pct':pct(v,total),'n_band':total})

# Cases
career_cases=[]
for x in linked:
    if x['dse_best5'] is not None and x['dse_best5']>=18 and x['level'] not in ['學士/本科']:
        career_cases.append({'grad_year':x['grad_year'],'code':x['anon'],'class_no':x.get('dse_class_no'), 'type':'高DSE但未入本科','metric':x['dse_best5'],'outcome':x['level'],'institution':x['institution'],'programme':x['programme'],'note':'可檢視JUPAS策略、面試/選科配對、學生志向或財政因素'})
    if x['dse_best5'] is not None and x['dse_best5']<=10 and x['level']=='學士/本科':
        career_cases.append({'grad_year':x['grad_year'],'code':x['anon'],'class_no':x.get('dse_class_no'), 'type':'低DSE但成功入本科','metric':x['dse_best5'],'outcome':x['level'],'institution':x['institution'],'programme':x['programme'],'note':'具升學策略成功個案價值，可分析課程選擇/面試/非成績因素'})
    if not x.get('meet3322') and x['level']=='學士/本科':
        career_cases.append({'grad_year':x['grad_year'],'code':x['anon'],'class_no':x.get('dse_class_no'), 'type':'未達3322但入本科/學位路徑','metric':x['dse_best5'],'outcome':x['level'],'institution':x['institution'],'programme':x['programme'],'note':'可作多元升學路徑案例，支援生涯規劃資訊'})
    if x['status']=='就業':
        career_cases.append({'grad_year':x['grad_year'],'code':x['anon'],'class_no':x.get('dse_class_no'), 'type':'直接就業','metric':x['dse_best5'],'outcome':x['industry'],'institution':'—','programme':'—','note':'建議跟進就業穩定性、行業培訓及日後進修需要'})
# career unlinked with important outcomes
for x in unlinked[:80]:
    if x['status'] in ['就業','其他','未知/未分類']:
        career_cases.append({'grad_year':x['grad_year'],'code':x['anon'],'class_no':x['class'], 'type':'出路需核實/未能連結DSE','metric':'—','outcome':x['status'],'institution':x['institution'],'programme':x['programme'],'note':'姓名未能唯一配對DSE，建議補學生ID/學號作精準追蹤'})
career_cases=career_cases[:250]

# Top institutions/programmes
career_top=[]
for dim in ['institution','programme','field','pathway']:
    for k,v in Counter(c[dim] for c in career if c[dim] and c[dim]!='未填').most_common(30):
        career_top.append({'dimension':dim,'item':k,'count':v})

# Update D
D['career']={'records':career,'linked_records':linked,'unlinked_count':len(unlinked),'dse_records':len(dse),'summary':career_summary,'outcome_by_dse':outcome_by_dse,'cases':career_cases,'top':career_top,'method_notes':[
    '已下載並分析2021-22至2024-25四份畢業生升學及就業調查Excel。',
    '以畢業年度及學生中文姓名將出路資料連結至同年DSE成績；輸出以匿名碼呈現。正式追蹤建議加入學生ID/STRN以減少同名或英文名填寫差異。',
    '「課程類別」「院校類型」「學習/職業範疇」為根據表單欄位及關鍵字標準化，個別課程需由升學及生涯規劃組再核實。'
]}
# update longitudinal career status
D.setdefault('longitudinal',{})['career_status']={'available':True,'note':f'已納入2022至2025畢業生出路資料，共{len(career)}筆回覆；其中{len(linked)}筆成功連結DSE成績，{len(unlinked)}筆需補學生ID/核實姓名。'}
# highlights
open(OUT,'w',encoding='utf-8').write(json.dumps(D,ensure_ascii=False,indent=2))
print('career',len(career),'linked',len(linked),'unlinked',len(unlinked),'cases',len(career_cases))
