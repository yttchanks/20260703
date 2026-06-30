import json, re, math, statistics, hashlib
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
BASE=Path('/home/user/drive_data')
OUT=Path('/home/user/analysis_output/analysis_data.json')
D=json.load(open(OUT,encoding='utf-8'))

def num(x):
    if x is None: return None
    if isinstance(x,(int,float)) and not isinstance(x,bool):
        return float(x) if not math.isnan(x) else None
    if isinstance(x,str):
        s=x.strip()
        if s.upper()=='U': return 0.0
        m=re.match(r'^(\d+(?:\.\d+)?)',s)
        return float(m.group(1)) if m else None
    return None

def norm_name(x):
    if x is None: return ''
    s=str(x).strip().upper()
    s=re.sub(r'\s+',' ',s)
    return s

def anon(*parts):
    raw='|'.join(map(lambda x:'' if x is None else str(x), parts))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:6].upper()

def stats(vals):
    vals=[v for v in vals if isinstance(v,(int,float))]
    if not vals: return {'n':0,'mean':None,'median':None}
    return {'n':len(vals),'mean':round(sum(vals)/len(vals),2),'median':round(statistics.median(vals),2)}

def corr(xs,ys):
    pairs=[(x,y) for x,y in zip(xs,ys) if isinstance(x,(int,float)) and isinstance(y,(int,float))]
    if len(pairs)<3: return None
    xs=[p[0] for p in pairs]; ys=[p[1] for p in pairs]
    mx=sum(xs)/len(xs); my=sum(ys)/len(ys)
    sx=math.sqrt(sum((x-mx)**2 for x in xs)); sy=math.sqrt(sum((y-my)**2 for y in ys))
    if sx==0 or sy==0: return None
    return round(sum((x-mx)*(y-my) for x,y in pairs)/(sx*sy),3)

# HKAT cohort students
hkat_students=[]
cohort_to_current={'2021-22':'S5','2022-23':'S4','2023-24':'S3','2024-25':'S2','2025-26':'S1'}
for f in sorted((BASE/'HKAT').glob('*_AT.xlsx')):
    cohort=f.name.replace('_AT.xlsx','')
    if cohort not in cohort_to_current: continue
    wb=load_workbook(f,data_only=True,read_only=False)
    if '總分' not in wb.sheetnames: continue
    ws=wb['總分']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        total=num(row[10] if len(row)>10 else None); avg=num(row[11] if len(row)>11 else None)
        if total is None: continue
        hkat_students.append({'cohort':cohort,'current_form_2025_26':cohort_to_current[cohort],'strn':str(row[1]),'ename':norm_name(row[2]),'cname':norm_name(row[3]),'sex':row[4],'hkat_math':num(row[7]),'hkat_eng':num(row[8]),'hkat_chi':num(row[9]),'hkat_total':total,'hkat_avg':avg,'anon':'L-'+anon(cohort,row[1],row[2],row[3])})
# index
idx_by_cname=defaultdict(list); idx_by_ename=defaultdict(list)
for s in hkat_students:
    if s['cname']: idx_by_cname[(s['cohort'],s['cname'])].append(s)
    if s['ename']: idx_by_ename[(s['cohort'],s['ename'])].append(s)

# Internal exam current 2025/26 extract, match relevant current form
exam_students=[]
for f in sorted((BASE/'Internal Exam/2025_26').glob('S[1-5].xlsx')):
    form=f.stem
    cohort={v:k for k,v in cohort_to_current.items()}.get(form)
    if not cohort: continue
    wb=load_workbook(f,data_only=True,read_only=False)
    for sheet in ['Exam1','Exam2','FT2']:
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]
        hr=None
        for r in range(1,min(15,ws.max_row)+1):
            vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            if '班別' in vals and '學生姓名' in vals:
                hr=r; break
        if not hr: continue
        headers=[ws.cell(hr,c).value for c in range(1,ws.max_column+1)]
        subj_cols=[(i,h) for i,h in enumerate(headers) if h and str(h) not in {'班別','號','學生姓名','總分','最高分','平均分','等級','班名次','級別名次','組別名次','科目組別','操行'} and '名次' not in str(h)]
        for row in ws.iter_rows(min_row=hr+1, values_only=True):
            if not row or row[0] is None or row[2] is None: continue
            cname=norm_name(row[2])
            avg=num(row[5] if len(row)>5 else None); rank=num(row[8] if len(row)>8 else None)
            fails=0; nsub=0
            for i,h in subj_cols:
                v=num(row[i] if i<len(row) else None)
                if v is not None:
                    nsub+=1
                    if v<50: fails+=1
            exam_students.append({'cohort':cohort,'form':form,'assessment':sheet,'class_no':f"{row[0]}#{int(row[1]) if isinstance(row[1],(int,float)) else row[1]}",'cname':cname,'exam_avg':avg,'rank':rank,'fails':fails,'n_subjects':nsub})

matched=[]; unmatched_exam=0
for e in exam_students:
    candidates=idx_by_cname.get((e['cohort'],e['cname']),[])
    if len(candidates)==1:
        s=candidates[0]
        matched.append({**e, 'anon':s['anon'], 'hkat_total':s['hkat_total'], 'hkat_avg':s['hkat_avg'], 'hkat_math':s['hkat_math'], 'hkat_eng':s['hkat_eng'], 'hkat_chi':s['hkat_chi'], 'sex':s['sex']})
    else:
        unmatched_exam+=1

# summarize by cohort and assessment
cohort_tracking=[]
for cohort, form in cohort_to_current.items():
    h=[s for s in hkat_students if s['cohort']==cohort]
    for assess in ['Exam1','FT2','Exam2']:
        m=[x for x in matched if x['cohort']==cohort and x['assessment']==assess]
        xs=[x['hkat_total'] for x in m]; ys=[x['exam_avg'] for x in m]
        cohort_tracking.append({'cohort':cohort,'current_form_2025_26':form,'stage':'入學→2025/26 '+assess,'hkat_n':len(h),'matched_n':len(m),'match_rate':round(100*len(m)/len(h),1) if h else None,'hkat_total_mean':stats([s['hkat_total'] for s in h])['mean'],'matched_hkat_mean':stats(xs)['mean'],'current_exam_avg_mean':stats(ys)['mean'],'hkat_exam_corr':corr(xs,ys),'multi_fail_3plus':sum(1 for x in m if x['fails']>=3),'multi_fail_rate':round(100*sum(1 for x in m if x['fails']>=3)/len(m),1) if m else None})

# longitudinal cases: admission low but improved, high potential at risk, persistent support
long_cases=[]
# percentiles by cohort/assessment
for cohort in cohort_to_current:
    for assess in ['Exam1','FT2','Exam2']:
        m=[x for x in matched if x['cohort']==cohort and x['assessment']==assess and x['exam_avg'] is not None]
        if len(m)<10: continue
        h_sorted=sorted([x['hkat_total'] for x in m]); e_sorted=sorted([x['exam_avg'] for x in m])
        hq10=h_sorted[max(0,int(len(h_sorted)*.1)-1)]; hq75=h_sorted[int(len(h_sorted)*.75)]; eq25=e_sorted[int(len(e_sorted)*.25)]; eq75=e_sorted[int(len(e_sorted)*.75)]
        for x in m:
            if x['hkat_total']<=hq10 and x['exam_avg']>=eq75:
                long_cases.append({'cohort':cohort,'form':x['form'],'assessment':assess,'code':x['anon'],'class_no':x['class_no'],'type':'低基線高進步','hkat_total':x['hkat_total'],'exam_avg':round(x['exam_avg'],1),'note':'入學基線偏低但現時考績達上四分位，可分析成功支援因素及作鼓勵個案'})
            elif x['hkat_total']>=hq75 and x['exam_avg']<=eq25:
                long_cases.append({'cohort':cohort,'form':x['form'],'assessment':assess,'code':x['anon'],'class_no':x['class_no'],'type':'高基線下滑風險','hkat_total':x['hkat_total'],'exam_avg':round(x['exam_avg'],1),'note':'入學基線較高但現時落入下四分位，需檢視動機、出席、科目適應或情緒因素'})
            elif x['hkat_total']<=hq10 and x['fails']>=3:
                long_cases.append({'cohort':cohort,'form':x['form'],'assessment':assess,'code':x['anon'],'class_no':x['class_no'],'type':'持續支援需要','hkat_total':x['hkat_total'],'exam_avg':round(x['exam_avg'],1),'note':'入學低基線且現時多科不合格，宜列入跨科個案會議優先處理'})
long_cases=long_cases[:200]

# DSE cohorts linkage availability and aggregate
# DSE 2025 entered S1 approx 2019-20, outside HKAT files. Use DSE aggregate and TSA cohort table as proxy.
dse_cohort_proxy=[]
for r in D.get('dse_passrate',[]):
    if 'year' in r and r.get('33222_pct') is not None:
        y=int(r['year']); entry=f"{y-6}-{str(y-5)[-2:]}"  # approximate Sept entry six years earlier
        dse_cohort_proxy.append({'dse_year':r['year'],'estimated_s1_entry':entry,'candidates':r.get('candidates'),'33222_pct':r.get('33222_pct'),'22222_pct':r.get('22222_pct'),'any5_2plus_pct':r.get('any5_2plus_pct'),'overall_2plus':r.get('overall_2plus')})

# progression/career: manifest unavailable
career_status={'available':False,'note':'Career Development資料夾內2021-22至2024-25畢業生升學及就業調查表均為失效捷徑（Original item deleted），未能建立升學/就業結果的實證追蹤。建議重新上載/分享原始回覆表，欄位至少包括畢業年度、匿名學生ID、DSE Best5/3322、去向類別、院校/課程/就業行業。'}

D['longitudinal']={'cohort_tracking':cohort_tracking,'longitudinal_cases':long_cases,'matched_records':len(matched),'hkat_students':len(hkat_students),'exam_records':len(exam_students),'unmatched_exam_records':unmatched_exam,'dse_cohort_proxy':dse_cohort_proxy,'career_status':career_status,'method_notes':[
    '以HKAT年度推算2025/26所在級別：2021-22→S5、2022-23→S4、2023-24→S3、2024-25→S2、2025-26→S1。',
    '現階段以中文姓名在HKAT與2025/26校內試作內部匹配，輸出只保留匿名碼；正式版本建議改用STRN/學校永久學生ID。',
    'DSE 2020-2025考生的S1入學年度早於現有HKAT檔案，因此只能做DSE cohort proxy，未能逐人由HKAT追到DSE。',
    '升學及就業調查表捷徑失效，暫未能將DSE成績連結至畢業去向。'
]}
# add high level highlights
D['highlights'].insert(0, {'area':'縱向追蹤', 'finding':f"已初步以姓名匿名匹配 HKAT→2025/26校內試：{len(matched)} 筆評估記錄，覆蓋 {len(hkat_students)} 名HKAT學生。", 'implication':'可建立入學基線到中三/高中校內試的增值與預警模型；DSE及升學/就業需補齊較早HKAT與Career資料。'})
open(OUT,'w',encoding='utf-8').write(json.dumps(D,ensure_ascii=False,indent=2))
print('longitudinal matched',len(matched),'cases',len(long_cases),'cohort rows',len(cohort_tracking))
