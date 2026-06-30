import json,re,math,hashlib,statistics
from pathlib import Path
from collections import defaultdict,Counter
from openpyxl import load_workbook, Workbook

BASE=Path('/home/user/drive_data')
ANALYSIS=Path('/home/user/analysis_output/analysis_data.json')
D=json.load(open(ANALYSIS,encoding='utf-8'))

def anon(*parts):
    return hashlib.sha1('|'.join(map(lambda x:'' if x is None else str(x), parts)).encode('utf-8')).hexdigest()[:6].upper()

def norm(x):
    if x is None: return ''
    return re.sub(r'\s+','',str(x).strip().upper())

def norm_name(x):
    if x is None: return ''
    return re.sub(r'\s+',' ',str(x).strip().upper())

def num(x):
    if x is None: return None
    if isinstance(x,(int,float)) and not isinstance(x,bool):
        return float(x) if not math.isnan(x) else None
    if isinstance(x,str):
        s=x.strip().upper()
        if s=='U': return 0.0
        if s=='A': return 'A'
        m=re.match(r'^(\d+(?:\.\d+)?)',s)
        return float(m.group(1)) if m else None
    return None

def pct(a,b): return round(100*a/b,1) if b else None

mapping=[]
def add(code,source,year='',cohort='',form='',klass='',no='',cname='',ename='',note=''):
    if not code: return
    mapping.append({'匿名碼':code,'來源':source,'年度':year,'入學cohort':cohort,'級別':form,'班別':klass,'學號':no,'中文姓名':cname or '', '英文姓名':ename or '', '備註':note})

# HKAT mappings: AT- and L-
cohort_to_current={'2021-22':'S5','2022-23':'S4','2023-24':'S3','2024-25':'S2','2025-26':'S1'}
for f in sorted((BASE/'HKAT').glob('*_AT.xlsx')):
    cohort=f.name.replace('_AT.xlsx','')
    wb=load_workbook(f,data_only=True,read_only=False)
    if '總分' not in wb.sheetnames: continue
    ws=wb['總分']
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row or row[0] is None: continue
        strn=row[1] if len(row)>1 else ''
        en=row[2] if len(row)>2 else ''
        cn=row[3] if len(row)>3 else ''
        sex=row[4] if len(row)>4 else ''
        add('AT-'+anon(strn,en),'HKAT個案',year=cohort,cohort=cohort,form=cohort_to_current.get(cohort,''),cname=cn,ename=en,note='HKAT匿名碼')
        add('L-'+anon(cohort,strn,en,cn),'縱向追蹤',year='2025/26',cohort=cohort,form=cohort_to_current.get(cohort,''),cname=cn,ename=en,note='HKAT→校內試追蹤碼')

# DSE mappings
def subject_columns(ws):
    header_row=None
    for r in range(1,min(6,ws.max_row)+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if '班別' in vals:
            header_row=r; break
    return header_row
for y in [2020,2021]:
    p=BASE/f'HKDSE/Reference/{y}/各班成績EXCEL'/('全部.xlsx' if y==2020 else '全部2021.xlsx')
    if not p.exists(): continue
    wb=load_workbook(p,data_only=True,read_only=False); ws=wb[wb.sheetnames[0]]; hr=subject_columns(ws)
    if not hr: continue
    for row in ws.iter_rows(min_row=hr+1,values_only=True):
        if not row or not row[0] or not row[1]: continue
        code='DSE-'+anon(str(y),row[0],row[1],row[2])
        add(code,'DSE個案',year=y,form='S6',klass=row[0],no=row[1],cname=row[3] if len(row)>3 else '',ename=row[2] if len(row)>2 else '',note='DSE匿名碼')
for y in [2022,2023,2024,2025]:
    p=BASE/f'HKDSE/各班成績/各班成績{y}.xlsx'
    if not p.exists(): p=BASE/f'HKDSE/Reference/{y}/各班成績{y}.xlsx'
    if not p.exists(): continue
    wb=load_workbook(p,data_only=True,read_only=False); ws=wb[wb.sheetnames[0]]; hr=subject_columns(ws)
    if not hr: continue
    for row in ws.iter_rows(min_row=hr+1,values_only=True):
        if not row or not row[0] or not row[1]: continue
        code='DSE-'+anon(str(y),row[0],row[1],row[2])
        add(code,'DSE個案',year=y,form='S6',klass=row[0],no=row[1],cname=row[3] if len(row)>3 else '',ename=row[2] if len(row)>2 else '',note='DSE匿名碼')

# Exam mappings EX-
for f in sorted((BASE/'Internal Exam/2025_26').glob('S[1-6].xlsx')):
    form=f.stem
    wb=load_workbook(f,data_only=True,read_only=False)
    for sheet in ['Exam1','Exam2','FT2']:
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]
        hr=None
        for r in range(1,min(15,ws.max_row)+1):
            vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            if '班別' in vals and '學生姓名' in vals:
                hr=r; break
        if not hr: continue
        for row in ws.iter_rows(min_row=hr+1,values_only=True):
            if not row or row[0] is None or row[2] is None: continue
            no=int(row[1]) if isinstance(row[1],(int,float)) else row[1]
            code='EX-'+anon(form,row[0],row[1],row[2])
            add(code,'校內試個案',year='2025/26',form=form,klass=row[0],no=no,cname=row[2],note=f'{sheet}匿名碼')

# Career mappings CD-
for f in sorted((BASE/'Career Deveolpment').glob('*.xlsx')):
    m=re.search(r'(20\d{2})-(20\d{2})',f.name)
    grad_year=int(m.group(2)) if m else ''
    wb=load_workbook(f,data_only=True,read_only=False)
    if '表單回應 1' not in wb.sheetnames: continue
    ws=wb['表單回應 1']
    for rid,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        if not row or not row[2]: continue
        code='CD-'+anon(grad_year,row[1],row[2])
        add(code,'升學就業個案',year=grad_year,form='S6畢業',klass=row[1],cname=row[2],note='出路匿名碼')

# Deduplicate mapping rows by code; if same code from multiple assessments merge notes
bycode={}
for r in mapping:
    c=r['匿名碼']
    if c not in bycode:
        bycode[c]=r
    else:
        # keep first but append note/source if different
        if r['備註'] not in bycode[c]['備註']:
            bycode[c]['備註'] += '；'+r['備註']
        if r['來源'] not in bycode[c]['來源']:
            bycode[c]['來源'] += ' / '+r['來源']
        for k in ['中文姓名','英文姓名','班別','學號','級別','年度','入學cohort']:
            if not bycode[c].get(k) and r.get(k): bycode[c][k]=r[k]
rows=list(bycode.values())
rows.sort(key=lambda r:(str(r['來源']),str(r['年度']),str(r['級別']),str(r['班別']),str(r['學號']),str(r['匿名碼'])))

# Create Excel mapping
wb=Workbook(); ws=wb.active; ws.title='匿名碼對照'
headers=['匿名碼','來源','年度','入學cohort','級別','班別','學號','中文姓名','英文姓名','備註']
ws.append(headers)
for r in rows:
    ws.append([r.get(h,'') for h in headers])
for col in range(1,len(headers)+1):
    ws.column_dimensions[chr(64+col)].width=18
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
# Legend
ws2=wb.create_sheet('使用說明')
ws2.append(['項目','說明'])
ws2.append(['用途','供校內授權人員把互動報告內的匿名碼對照至真實學生姓名作跟進。'])
ws2.append(['不包括','沒有包含電話號碼等非必要個人資料。'])
ws2.append(['保密','請勿把本檔與公開/會議版匿名報告一併廣泛分享。'])
ws2.append(['代碼類型','AT=HKAT；L=縱向追蹤；DSE=DSE個案；EX=校內試；CD=升學就業。'])
map_path=Path('/home/user/anonymous_code_name_mapping.xlsx')
wb.save(map_path)

# Audit checks
checks=[]
def check(name, ok, detail):
    checks.append({'檢查項目':name,'結果':'通過' if ok else '需檢視','詳情':detail})
# Career source rows
career_files=list((BASE/'Career Deveolpment').glob('*.xlsx'))
source_count=0
for f in career_files:
    wb2=load_workbook(f,data_only=True,read_only=False); wsx=wb2['表單回應 1']
    cnt=sum(1 for row in wsx.iter_rows(min_row=2,values_only=True) if row and row[2])
    source_count+=cnt
check('畢業生出路來源筆數', source_count==len(D.get('career',{}).get('records',[])), f"Excel有效回覆 {source_count}；分析資料 {len(D.get('career',{}).get('records',[]))}")
# Linked count not exceed records
C=D.get('career',{})
check('DSE與出路連結筆數', len(C.get('linked_records',[]))<=len(C.get('records',[])), f"連結 {len(C.get('linked_records',[]))} / 出路 {len(C.get('records',[]))}")
# Career summary percent totals per year/dim approximately 100
bad=[]
for (y,dim),arr in defaultdict(list, {}).items(): pass
groups=defaultdict(list)
for r in C.get('summary',[]): groups[(r['grad_year'],r['dimension'])].append(r)
for (y,dim),arr in groups.items():
    s=sum(r['pct'] for r in arr if isinstance(r.get('pct'),(int,float)))
    if not (99.5<=s<=100.5): bad.append((y,dim,s))
check('出路摘要百分比', not bad, '所有年度/維度百分比合計約100%' if not bad else str(bad[:5]))
# Case codes are mappable
case_codes=set()
for c in D.get('hkat_cases',[]): case_codes.add(c.get('code'))
for c in D.get('dse_cases',[]): case_codes.add(c.get('code'))
for c in D.get('exam_cases',[]): case_codes.add(c.get('code'))
for c in D.get('longitudinal',{}).get('longitudinal_cases',[]): case_codes.add(c.get('code'))
for c in C.get('cases',[]): case_codes.add(c.get('code'))
missing=[c for c in case_codes if c and c not in bycode]
check('個案匿名碼對照覆蓋', not missing, f"個案碼 {len(case_codes)} 個；未能對照 {len(missing)} 個")
# Duplicate codes with conflicting names
conflicts=[]
raw=defaultdict(set)
for r in mapping:
    raw[r['匿名碼']].add((r['中文姓名'],r['英文姓名']))
for code,names in raw.items():
    if len(names)>1: conflicts.append((code,list(names)[:3]))
check('匿名碼唯一性', not conflicts, '沒有同一匿名碼對應多個姓名' if not conflicts else str(conflicts[:5]))
# No names in enhanced report (after later sanitize maybe currently can check after generation)

# Write audit workbook
awb=Workbook(); a=awb.active; a.title='數據檢查'
a.append(['檢查項目','結果','詳情'])
for r in checks: a.append([r['檢查項目'],r['結果'],r['詳情']])
a2=awb.create_sheet('核心數量')
a2.append(['項目','數量'])
a2.append(['HKAT摘要列',len(D.get('hkat',[]))]); a2.append(['TSA摘要列',len(D.get('tsa',[]))]); a2.append(['DSE科目摘要列',len(D.get('dse_subject',[]))]); a2.append(['校內試摘要列',len(D.get('exam',[]))]); a2.append(['出路回覆',len(C.get('records',[]))]); a2.append(['出路連結DSE',len(C.get('linked_records',[]))]); a2.append(['匿名碼對照數',len(rows)])
a3=awb.create_sheet('出路年度統計')
a3.append(['年度','回覆','升學','就業','重讀','本科/學士','副學士','高級文憑','文憑/職專/DAE'])
for y in sorted(set(r['grad_year'] for r in C.get('records',[]))):
    arr=[r for r in C['records'] if r['grad_year']==y]
    a3.append([y,len(arr),sum(r['status']=='升學' for r in arr),sum(r['status']=='就業' for r in arr),sum(r['status']=='重讀' for r in arr),sum(r['level']=='學士/本科' for r in arr),sum(r['level']=='副學士' for r in arr),sum(r['level']=='高級文憑' for r in arr),sum(r['level'] in ['文憑/基礎/職專','應用教育/毅進'] for r in arr)])
for sheet in awb.worksheets:
    sheet.freeze_panes='A2'; sheet.auto_filter.ref=sheet.dimensions
    for col in range(1, min(sheet.max_column,10)+1): sheet.column_dimensions[chr(64+col)].width=22
audit_path=Path('/home/user/data_audit_summary.xlsx')
awb.save(audit_path)
print('mapping',map_path, len(rows),'audit',audit_path, checks)
