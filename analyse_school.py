import os, re, json, math, statistics, glob, hashlib
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import load_workbook

BASE=Path('/home/user/drive_data')
OUT=Path('/home/user/analysis_output'); OUT.mkdir(exist_ok=True)

def num(x):
    if x is None: return None
    if isinstance(x,(int,float)) and not isinstance(x,bool):
        if math.isnan(x): return None
        return float(x)
    if isinstance(x,str):
        s=x.strip()
        if s=='' or s in ['-','--','N.A.','NA','Abs','ABS','缺席']: return None
        if s.upper()=='U': return 0.0
        if s.upper()=='A': return 1.0 # attained, not score
        m=re.match(r'^(\d+(?:\.\d+)?)',s)
        if m: return float(m.group(1))
    return None

def lvl(x):
    if x is None: return None
    if isinstance(x,(int,float)): return float(x)
    if isinstance(x,str):
        s=x.strip().upper()
        if s in ('U','UNCL','ABS','-',''): return 0.0 if s=='U' else None
        if s=='A': return 'A'
        m=re.match(r'^(\d+)',s)
        if m: return float(m.group(1))
    return None

def stats(vals):
    vals=[v for v in vals if isinstance(v,(int,float)) and not math.isnan(v)]
    if not vals: return {'n':0,'mean':None,'median':None,'min':None,'max':None,'sd':None}
    return {'n':len(vals),'mean':round(sum(vals)/len(vals),2),'median':round(statistics.median(vals),2),'min':round(min(vals),2),'max':round(max(vals),2),'sd':round(statistics.pstdev(vals),2) if len(vals)>1 else 0}

def pct(a,b): return round(100*a/b,1) if b else None

def anon(*parts):
    raw='|'.join(map(lambda x:'' if x is None else str(x), parts))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:6].upper()

def year_from_path(p):
    s=str(p)
    m=re.search(r'(20\d{2})[-_/ ]?(?:-|/)?(\d{2})?',s)
    return m.group(1) if m else None

# HKAT
hkat=[]; hkat_cases=[]
for f in sorted((BASE/'HKAT').glob('*_AT.xlsx')):
    yr=f.name.replace('_AT.xlsx','')
    wb=load_workbook(f,data_only=True,read_only=False)
    # subject summaries
    student_scores={}
    for subj in ['中文','英文','數學']:
        if subj not in wb.sheetnames: continue
        ws=wb[subj]; header=[c.value for c in ws[1]]
        # find total col: last named 總分 or for English maybe Total
        total_idx=None
        for i,h in enumerate(header):
            if h and str(h).strip().lower() in ['總分','total']:
                total_idx=i; break
        if total_idx is None: total_idx=ws.max_column-1
        vals=[]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            v=num(row[total_idx] if total_idx<len(row) else None)
            if v is None: continue
            vals.append(v)
            sid=str(row[1]) if len(row)>1 else str(row[0])
            student_scores.setdefault(sid, {'id':sid,'name':row[2] if len(row)>2 else '', 'cname':row[3] if len(row)>3 else '', 'sex':row[4] if len(row)>4 else '', 'scores':{}})['scores'][subj]=v
        st=stats(vals); st.update({'year':yr,'subject':subj})
        hkat.append(st)
    # cases from 總分 sheet if exists
    if '總分' in wb.sheetnames:
        ws=wb['總分']; header=[c.value for c in ws[1]]
        # expected columns
        rows=[]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            mathv=num(row[7] if len(row)>7 else None); eng=num(row[8] if len(row)>8 else None); ch=num(row[9] if len(row)>9 else None); total=num(row[10] if len(row)>10 else None); avg=num(row[11] if len(row)>11 else None)
            if total is None: continue
            rows.append({'year':yr,'code':'AT-'+anon(row[1],row[2]),'sex':row[4] if len(row)>4 else '', 'math':mathv,'english':eng,'chinese':ch,'total':total,'avg':avg})
        if rows:
            totals=[r['total'] for r in rows]
            q10=sorted(totals)[max(0,int(len(totals)*0.1)-1)]
            q90=sorted(totals)[min(len(totals)-1,int(len(totals)*0.9))]
            for r in rows:
                scores=[r['math'],r['english'],r['chinese']]
                if all(v is not None for v in scores):
                    spread=max(scores)-min(scores)
                    if r['total']<=q10:
                        hkat_cases.append({**r,'type':'入學測驗總分最低10%','note':'需要及早配對中英數基礎支援'})
                    elif r['total']>=q90:
                        hkat_cases.append({**r,'type':'入學測驗高潛能首10%','note':'可安排拔尖/領袖/競賽延展'})
                    elif spread>=45:
                        hkat_cases.append({**r,'type':'三科表現懸殊','note':'強弱科差距大，宜個別化學習計劃'})
# limit cases by year/type
hkat_cases=sorted(hkat_cases,key=lambda x:(x['year'],x['type'], x.get('total') or 0))[:120]

# DSE student-level and subject-level from all-class files (avoid duplicates by using HKDSE/各班成績 and some reference)
dse_files=[]
for y in [2020,2021]:
    dse_files.append((str(y), BASE/f'HKDSE/Reference/{y}/各班成績EXCEL'/('全部.xlsx' if y==2020 else '全部2021.xlsx')))
for y in [2022,2023,2024,2025]:
    p=BASE/f'HKDSE/各班成績/各班成績{y}.xlsx'
    if not p.exists(): p=BASE/f'HKDSE/Reference/{y}/各班成績{y}.xlsx'
    dse_files.append((str(y),p))

def subject_columns(ws):
    # locate header row with 班別
    header_row=None
    for r in range(1,min(6,ws.max_row)+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if '班別' in vals:
            header_row=r; break
    if not header_row: return None,None,[]
    hdr=[ws.cell(header_row,c).value for c in range(1,ws.max_column+1)]
    grp=[ws.cell(header_row-1,c).value if header_row>1 else None for c in range(1,ws.max_column+1)]
    cols=[]
    for i,h in enumerate(hdr):
        hstr=str(h).strip() if h is not None else ''
        gstr=str(grp[i]).strip() if grp[i] is not None else ''
        subj=None
        if gstr.upper()=='CHINESE' and hstr in ['整體','Overall']:
            subj='中文'
        elif gstr.upper()=='ENGLISH' and hstr in ['Overall','整體','TOTAL']:
            subj='英文'
        elif hstr in ['數學','公民','M1','M2','會計','商管','生物','化學','物理','中史','歷史','地理','旅款','經濟','ICT','BAFS','中國文學','視藝','企會財','企管','倫理與宗教']:
            subj=hstr
        if subj:
            cols.append((i,subj))
    return header_row,hdr,cols

dse_subject=[]; dse_students=[]; dse_cases=[]
for yr,p in dse_files:
    if not p.exists(): continue
    wb=load_workbook(p,data_only=True,read_only=False)
    ws=wb[wb.sheetnames[0]]
    hr,hdr,cols=subject_columns(ws)
    if not cols: continue
    subvals=defaultdict(list); entries=[]
    for row in ws.iter_rows(min_row=hr+1, values_only=True):
        if not row or not row[0] or str(row[0]).startswith('Note'): continue
        sid=f"{row[0]}-{row[1]}"
        scores={}
        for idx,subj in cols:
            v=lvl(row[idx] if idx<len(row) else None)
            if isinstance(v,(int,float)):
                scores[subj]=v; subvals[subj].append(v)
            elif v=='A':
                scores[subj]='A'
        if not scores: continue
        numeric=[v for v in scores.values() if isinstance(v,(int,float))]
        best5=sum(sorted(numeric, reverse=True)[:5]) if len(numeric)>=5 else None
        meet3322 = (scores.get('中文',0)>=3 if isinstance(scores.get('中文'),(int,float)) else False) and (scores.get('英文',0)>=3 if isinstance(scores.get('英文'),(int,float)) else False) and (scores.get('數學',0)>=2 if isinstance(scores.get('數學'),(int,float)) else False) and sum(1 for k,v in scores.items() if k not in ['中文','英文','數學'] and isinstance(v,(int,float)) and v>=2)>=2
        entry={'year':yr,'code':f"{row[0]}#{int(row[1]) if isinstance(row[1],(int,float)) else row[1]}",'anon':'DSE-'+anon(yr,row[0],row[1],row[2]),'class':row[0],'no':row[1],'scores':scores,'best5':best5,'meet3322':meet3322}
        entries.append(entry); dse_students.append(entry)
    for subj,vals in subvals.items():
        n=len(vals)
        s=stats(vals)
        dse_subject.append({'year':yr,'subject':subj,**s,'lv2_plus':pct(sum(v>=2 for v in vals),n),'lv3_plus':pct(sum(v>=3 for v in vals),n),'lv4_plus':pct(sum(v>=4 for v in vals),n),'u_rate':pct(sum(v==0 for v in vals),n)})
    # cases
    if entries:
        for e in sorted(entries, key=lambda x: (x['best5'] if x['best5'] is not None else -1), reverse=True)[:8]:
            dse_cases.append({'year':yr,'code':e['anon'],'class_no':e['code'],'type':'DSE高增值/高成就','metric':e['best5'],'note':'Best 5 最高群，可作拔尖及升學品牌個案（內部查核姓名）'})
        for e in sorted([x for x in entries if x['best5'] is not None], key=lambda x:x['best5'])[:8]:
            dse_cases.append({'year':yr,'code':e['anon'],'class_no':e['code'],'type':'DSE支援需要','metric':e['best5'],'note':'Best 5 偏低或核心科未達標，需檢視中四至中六支援軌跡'})
        for e in entries:
            sc=e['scores']
            if isinstance(sc.get('中文'),(int,float)) and isinstance(sc.get('英文'),(int,float)) and isinstance(sc.get('數學'),(int,float)):
                if (sc.get('中文')==2 and sc.get('英文')>=3 and sc.get('數學')>=2) or (sc.get('英文')==2 and sc.get('中文')>=3 and sc.get('數學')>=2):
                    dse_cases.append({'year':yr,'code':e['anon'],'class_no':e['code'],'type':'接近3322邊緣','metric':e['best5'],'note':'一科語文只差一級，適合重點卷別診斷及操練'})
# limit
# keep recent and representative
dse_cases=dse_cases[:200]

# DSE historical pass-rate workbook
passrate=[]
prfile=BASE/'HKDSE/Reference/2024/DSE Passing Rate 2012-2024(每年填寫)_更新於240718.xlsx'
if prfile.exists():
    wb=load_workbook(prfile,data_only=True,read_only=False)
    if '達標統計' in wb.sheetnames:
        ws=wb['達標統計']
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]: continue
            m=re.search(r'(20\d{2})',str(row[0]))
            if m:
                passrate.append({'year':m.group(1),'33222_pct':pct(row[2],1) if isinstance(row[2],(int,float)) and row[2]<=1 else row[2], '22222_pct':pct(row[6],1) if isinstance(row[6],(int,float)) and row[6]<=1 else row[6], 'any5_2plus_pct':pct(row[8],1) if isinstance(row[8],(int,float)) and row[8]<=1 else row[8], 'overall_2plus':pct(row[9],1) if isinstance(row[9],(int,float)) and row[9]<=1 else row[9], 'candidates':row[11]})
    if '各科2級或以上' in wb.sheetnames:
        ws=wb['各科2級或以上']
        headers=[c.value for c in ws[1]]
        for r in range(2,ws.max_row+1,2):
            yr=ws.cell(r,1).value
            if not isinstance(yr,(int,float)): continue
            for c in range(2, min(ws.max_column,23)+1):
                subj=headers[c-1]
                val=num(ws.cell(r,c).value)
                if subj and val is not None:
                    passrate.append({'year':str(int(yr)),'subject':str(subj),'metric':'2plus_subject_pct','value':round(val,1)})

# Nine-value table
nine=[]
nfile=BASE/'HKDSE/Reference/HKDSE_九位數表_20251024.xlsx'
if nfile.exists():
    wb=load_workbook(nfile,data_only=True,read_only=False)
    ws=wb['九位數']
    years=[]
    for c in range(2,ws.max_column+1):
        m=re.search(r'(20\d{2})',str(ws.cell(1,c).value)) if ws.cell(1,c).value else None
        years.append(m.group(1) if m else None)
    for r in range(4,ws.max_row+1):
        subj=ws.cell(r,1).value
        if not subj: continue
        for idx,y in enumerate(years,start=2):
            v=num(ws.cell(r,idx).value)
            if y and v is not None: nine.append({'year':y,'subject':str(subj),'nine_value':v})

# Internal Exam (2025/26 main S workbooks; latest Exam1/Exam2)
exam=[]; exam_cases=[]
for f in sorted((BASE/'Internal Exam/2025_26').glob('S[1-6].xlsx')):
    form=f.stem
    wb=load_workbook(f,data_only=True,read_only=False)
    for sheet in ['Exam1','Exam2','FT2']:
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]
        hr=None
        for r in range(1, min(15,ws.max_row)+1):
            vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            if '班別' in vals and '學生姓名' in vals:
                hr=r; break
        if not hr: continue
        headers=[ws.cell(hr,c).value for c in range(1,ws.max_column+1)]
        skip={'班別','號','學生姓名','總分','最高分','平均分','等級','班名次','級別名次','組別名次','科目組別','操行'}
        subj_cols=[(i,h) for i,h in enumerate(headers) if h and str(h).strip() not in skip and not str(h).strip().endswith('名次')]
        subvals=defaultdict(list); students=[]
        for row in ws.iter_rows(min_row=hr+1,values_only=True):
            if not row or row[0] is None or row[2] is None: continue
            avg=num(row[5] if len(row)>5 else None); rank=num(row[8] if len(row)>8 else None)
            fails=0; scores={}
            for idx,h in subj_cols:
                v=num(row[idx] if idx<len(row) else None)
                if v is not None:
                    scores[str(h)]=v; subvals[str(h)].append(v)
                    if v<50: fails+=1
            students.append({'code':f"{row[0]}#{int(row[1]) if isinstance(row[1],(int,float)) else row[1]}", 'anon':'EX-'+anon(form,row[0],row[1],row[2]),'avg':avg,'rank':rank,'fails':fails,'scores':scores})
        for subj,vals in subvals.items():
            st=stats(vals)
            exam.append({'year':'2025/26','form':form,'assessment':sheet,'subject':subj,**st,'pass_rate':pct(sum(v>=50 for v in vals),len(vals))})
        if students:
            for e in sorted([x for x in students if x['avg'] is not None], key=lambda x:x['avg'])[:10]:
                exam_cases.append({'form':form,'assessment':sheet,'code':e['anon'],'class_no':e['code'],'type':'校內試整體低分','metric':round(e['avg'],1),'note':'整體平均較低，宜跟進出席、功課、測考及家校支援'})
            for e in [x for x in students if x['fails']>=3][:20]:
                exam_cases.append({'form':form,'assessment':sheet,'code':e['anon'],'class_no':e['code'],'type':'多科不合格','metric':e['fails'],'note':'三科或以上低於50，需跨科支援會議'})
# rank movement between Exam1 and Exam2 for each form if available
# skipped if identifiers insufficient
exam_cases=exam_cases[:250]

# TSA
tsa=[]
tsafile=BASE/'TSA/TSA_中英數總表_20251123.xlsx'
if tsafile.exists():
    wb=load_workbook(tsafile,data_only=True,read_only=False)
    if '報告' in wb.sheetnames:
        ws=wb['報告']
        # fixed blocks: Chi cols 12-17, Eng 19-24, Math 26-31 (1-based approximate)
        blocks=[('中國語文',12),('英文',19),('數學',26)]
        for row in ws.iter_rows(min_row=3, max_row=60, values_only=True):
            if not row or row[0] is None: continue
            m=re.search(r'(20\d{2})',str(row[0]))
            if not m: continue
            year=m.group(1)
            for subj,start in blocks:
                A=num(row[start] if len(row)>start else None)      # zero-based? start is 12 -> col13 A
                B=num(row[start+1] if len(row)>start+1 else None)
                school=num(row[start+2] if len(row)>start+2 else None)
                hk=num(row[start+3] if len(row)>start+3 else None)
                dse=num(row[start+4] if len(row)>start+4 else None)
                va=num(row[start+5] if len(row)>start+5 else None)
                # for Chinese actual row cols are: HKAT col11, A12, B13, school14, hk15, dse16, va17 (0-based 11-17)
                # adjust by subject
                if subj=='中國語文':
                    A=num(row[12]) if len(row)>12 else None; B=num(row[13]) if len(row)>13 else None; school=num(row[14]) if len(row)>14 else None; hk=num(row[15]) if len(row)>15 else None; dse=num(row[16]) if len(row)>16 else None; va=num(row[17]) if len(row)>17 else None
                elif subj=='英文':
                    A=num(row[19]) if len(row)>19 else None; B=num(row[20]) if len(row)>20 else None; school=num(row[21]) if len(row)>21 else None; hk=num(row[22]) if len(row)>22 else None; dse=num(row[23]) if len(row)>23 else None; va=num(row[24]) if len(row)>24 else None
                else:
                    A=num(row[26]) if len(row)>26 else None; B=num(row[27]) if len(row)>27 else None; school=num(row[28]) if len(row)>28 else None; hk=num(row[29]) if len(row)>29 else None; dse=num(row[30]) if len(row)>30 else None; va=num(row[31]) if len(row)>31 else None
                if school is not None:
                    tsa.append({'year':year,'subject':subj,'n':A,'basic_n':B,'school_basic_pct':school,'hk_basic_pct':hk,'dse_pass_pct':dse,'value_added':va})
    if '報告 SEN' in wb.sheetnames:
        ws=wb['報告 SEN']
        blocks=[('中國語文',1),('英文',5),('數學',9)]
        for row in ws.iter_rows(min_row=3, max_row=20, values_only=True):
            if not row or row[0] is None: continue
            m=re.search(r'(20\d{2})',str(row[0]))
            if not m: continue
            for subj,start in blocks:
                school=num(row[start+2]) if len(row)>start+2 else None
                hk=num(row[start+3]) if len(row)>start+3 else None
                if school is not None:
                    tsa.append({'year':m.group(1),'subject':subj+'(SEN)','n':num(row[start]),'basic_n':num(row[start+1]),'school_basic_pct':school,'hk_basic_pct':hk})

# Department reports synthesis input
# Map departments by subject group
subject_groups={
    '語文教育':['中文','中國語文','英文','Ch Lang','Eng Lang','CHINESE','ENGLISH'],
    '數學教育':['數學','Maths','M1','M2'],
    '科學教育':['生物','化學','物理','科學','科初','Bio','Chem','Phy'],
    '人文及社會科學':['中史','歷史','地理','公民','公經社','經濟','LS','Ch Hist','Hist','Geog','Econ'],
    '商業及科技':['會計','商管','BAFS','BUS','ACCT','ICT','企會財','企管'],
    '升學及生涯規劃':['DSE高成就','DSE支援','畢業生升學'],
    '學習支援/SEN':['SEN','多科不合格','低分','Low achiever'],
}

summary={
    'generated_date':'2026-06-29',
    'source':'Google Drive folder: 20260703 Staff Development',
    'manifest_counts':{},
}
try:
    manifest=json.load(open('/home/user/drive_manifest.json',encoding='utf-8'))
    summary['manifest_counts']=dict(Counter(r['type'] for r in manifest))
    summary['top_folders']=dict(Counter(r['path'].split('/')[1] if '/' in r['path'] else r['path'] for r in manifest))
except Exception: pass

# compute narrative highlights
highlights=[]
def last_two(records, subject_key='subject', value_key='mean', group=None):
    pass
# HKAT trend highlights
for subj in ['中文','英文','數學']:
    arr=sorted([r for r in hkat if r['subject']==subj and r['mean'] is not None], key=lambda x:x['year'])
    if len(arr)>=2:
        diff=round(arr[-1]['mean']-arr[0]['mean'],2)
        highlights.append({'area':'HKAT '+subj,'finding':f"{arr[0]['year']}至{arr[-1]['year']}平均分變化 {diff:+.2f}（{arr[0]['mean']}→{arr[-1]['mean']}）",'implication':'可作中一入學基線及分層支援指標'})
# DSE recent subject low/high
recent=max([r['year'] for r in dse_subject], default=None)
if recent:
    recent_sub=[r for r in dse_subject if r['year']==recent and r.get('lv2_plus') is not None]
    for r in sorted(recent_sub,key=lambda x:x['lv2_plus'])[:5]:
        highlights.append({'area':'HKDSE '+r['subject'],'finding':f"{recent} Level 2+：{r['lv2_plus']}%（n={r['n']}）",'implication':'屬近期較需關注科目/組別，建議進一步按班別及卷別檢視'})
# Exam recent lows
if exam:
    for r in sorted([x for x in exam if x['assessment']=='Exam2' and x.get('pass_rate') is not None], key=lambda x:x['pass_rate'])[:8]:
        highlights.append({'area':f"校內試 {r['form']} {r['subject']}", 'finding':f"Exam2 合格率 {r['pass_rate']}%，平均 {r['mean']}", 'implication':'建議科組檢視評估設計、錯題類型及補底安排'})
# TSA gap
for r in sorted([x for x in tsa if x.get('hk_basic_pct') is not None and x.get('school_basic_pct') is not None], key=lambda x:(x['school_basic_pct']-x['hk_basic_pct']))[:6]:
    highlights.append({'area':'TSA '+r['subject'], 'finding':f"{r['year']} 學校基本水平 {r['school_basic_pct']}%，全港 {r['hk_basic_pct']}%（差距 {round(r['school_basic_pct']-r['hk_basic_pct'],1):+.1f}）", 'implication':'作全校語文/數學基礎能力及SEN支援參考'})

# SWOT / Vision templates with data hooks per group
swot={
    '語文教育':{
        'vision':'以讀寫聽說數據診斷推動跨級語文能力階梯，提升DSE核心語文達標及高階表達。',
        'S':['已有HKAT、TSA、DSE及校內試多源數據，可追蹤入學基線至公開試。'],
        'W':['英文/TSA及DSE核心語文邊緣個案需更早識別；卷別資料未完全轉化為恆常教學行動。'],
        'O':['可建立S1入學至S6 DSE的語文能力檔案，配合閱讀、寫作及口語分層任務。'],
        'T':['學生入學差異、低動機及SEN比例上升會放大核心語文達標風險。']
    },
    '數學教育':{
        'vision':'建立由HKAT數學分項、TSA基本水平、校內試至DSE的縱向增值模型。',
        'S':['數學在HKAT及校內試具清晰分項/總分，可快速定位代數、度量、數據處理弱點。'],
        'W':['部分級別合格率及低分群集中，需將補底由考後改為考前預警。'],
        'O':['可推行錯題庫、概念地圖及小步子重測；M1/M2可作拔尖路徑。'],
        'T':['基礎斷層會影響高中選科及DSE 3322/升學門檻。']
    },
    '科學教育':{
        'vision':'以探究與數據素養提升科學科DSE Level 2+/4+及STEM學習動機。',
        'S':['DSE理科選修科有多年成績，可識別高成就及弱項科目。'],
        'W':['選修人數相對小，單年波動大，需合併多年趨勢判讀。'],
        'O':['可透過跨科STEM、實驗技能rubric及早期選科輔導提升留科質素。'],
        'T':['高中選科人數下跌或基礎數學弱會限制理科表現。']
    },
    '人文及社會科學':{
        'vision':'以閱讀、資料詮釋及論述能力為核心，提升人文科目穩定達標與高階思維。',
        'S':['中史、歷史、地理、公民等可與語文能力及DSE達標數據互相參照。'],
        'W':['不同科目評核形式差異大，統一數據口徑仍需加強。'],
        'O':['可建立資料題/論述題共通rubric，跨科分享高效題型教學。'],
        'T':['閱讀耐性及時事知識不足會拖低資料回應題表現。']
    },
    '商業及科技':{
        'vision':'發展商科及科技科的應用學習、數據分析及職涯連結，提高選修吸引力與成績穩定性。',
        'S':['DSE商管、會計、ICT等有可比較的歷年Level 2+/4+。'],
        'W':['選修科樣本較小，須避免只看單年百分比。'],
        'O':['可連結生涯規劃、企業案例、AI/數據應用作課程亮點。'],
        'T':['學生基礎數學/語文弱會限制案例分析及計算題表現。']
    },
    '學習支援/SEN':{
        'vision':'以早識別、早介入、跨科個案會議，建立可量度的支援成效循環。',
        'S':['TSA SEN、校內試多科不合格及HKAT低分群提供具體名單線索（本報告以匿名碼呈現）。'],
        'W':['資料分散於多個系統/檔案，需整合個案追蹤表。'],
        'O':['可用預警儀表板每次考績後自動更新支援優先次序。'],
        'T':['支援滯後會令低成就由單科擴展至跨科，影響出席及升級。']
    },
    '升學及生涯規劃':{
        'vision':'用DSE Best5、3322、選修科表現與畢業出路數據，形成選科—備試—升學一體化支援。',
        'S':['DSE多年數據可支援升學門檻、拔尖補底及選科諮詢。'],
        'W':['Career Development表單捷徑顯示原項目已刪除，現時未能納入畢業生出路趨勢。'],
        'O':['可重建畢業生出路資料庫，與DSE成績及選科組合連結分析。'],
        'T':['若出路數據缺失，年度檢討較難評估課程與升學支援成效。']
    }
}

# data quality notes
quality=[]
if summary['manifest_counts'].get('unavailable'):
    quality.append(f"Career Development內有 {summary['manifest_counts'].get('unavailable')} 個Google Sheets捷徑顯示原項目已刪除，未能分析畢業生出路。")
quality.append('本輸出已將學生個案匿名化；如要內部跟進，可用原Excel按班別/學號對照。')
quality.append('PDF報表數量多，今次主要以可結構化的Excel/文字檔作量化分析；PDF可作下一輪OCR/文本摘要補充。')

analysis={
    'summary':summary,
    'highlights':highlights,
    'hkat':hkat,
    'hkat_cases':hkat_cases,
    'dse_subject':dse_subject,
    'dse_cases':dse_cases,
    'dse_students_count':len(dse_students),
    'dse_passrate':passrate,
    'nine_value':nine,
    'exam':exam,
    'exam_cases':exam_cases,
    'tsa':tsa,
    'swot':swot,
    'subject_groups':subject_groups,
    'quality_notes':quality,
}
open(OUT/'analysis_data.json','w',encoding='utf-8').write(json.dumps(analysis,ensure_ascii=False,indent=2))
print('HKAT',len(hkat),len(hkat_cases),'DSE subj',len(dse_subject),'DSE cases',len(dse_cases),'Exam',len(exam),'Exam cases',len(exam_cases),'TSA',len(tsa),'Passrate',len(passrate),'Nine',len(nine))
