import json, re, math, statistics
from pathlib import Path
from collections import Counter,defaultdict
from openpyxl import load_workbook
D=json.load(open('/home/user/analysis_output/analysis_data.json',encoding='utf-8'))

def pct(a,b): return round(100*a/b,1) if b else None

def detail_level(r):
    level=r.get('level') or r.get('課程層級') or ''
    if level=='學士/本科':
        pathway=(r.get('pathway') or '')+' '+(r.get('institution_type') or '')+' '+(r.get('institution') or '')
        if any(k.lower() in pathway.lower() for k in ['jupas','聯招','sssdP'.lower(),'教資會','指定專業']):
            return '學位課程（JUPAS/SSSDP）'
        return '學位課程（Non-JUPAS/自資/其他）'
    return level or '未分類'

C=D.get('career',{})
for key in ['records','linked_records']:
    for r in C.get(key,[]):
        r['level_detail']=detail_level(r)
# summary for detailed level by year
# Remove old level_detail rows if any
C['summary']=[r for r in C.get('summary',[]) if r.get('dimension')!='level_detail']
for y in sorted(set(r.get('grad_year') for r in C.get('records',[]) if r.get('grad_year'))):
    arr=[r for r in C['records'] if r.get('grad_year')==y]
    total=len(arr)
    for k,v in Counter(r.get('level_detail','未分類') for r in arr).most_common():
        C['summary'].append({'grad_year':y,'dimension':'level_detail','category':k,'count':v,'pct':pct(v,total)})
# outcome_by_dse for detailed level
C['outcome_by_dse']=[r for r in C.get('outcome_by_dse',[]) if r.get('dimension')!='level_detail']
bands=[('0-9',0,9),('10-14',10,14),('15-19',15,19),('20+',20,999)]
for name,lo,hi in bands:
    arr=[x for x in C.get('linked_records',[]) if x.get('dse_best5') is not None and lo<=x.get('dse_best5')<=hi]
    total=len(arr)
    for k,v in Counter(x.get('level_detail','未分類') for x in arr).most_common():
        C['outcome_by_dse'].append({'best5_band':name,'dimension':'level_detail','category':k,'count':v,'pct':pct(v,total),'n_band':total})
for flag,label in [(True,'3322達標'),(False,'3322未達標')]:
    arr=[x for x in C.get('linked_records',[]) if x.get('meet3322')==flag]
    total=len(arr)
    for k,v in Counter(x.get('level_detail','未分類') for x in arr).most_common():
        C['outcome_by_dse'].append({'best5_band':label,'dimension':'level_detail','category':k,'count':v,'pct':pct(v,total),'n_band':total})
D['career']=C

# Promotion/retention line
prom_file=Path('/home/user/drive_data/留班線/升留班線.xlsx')
promotion=[]
if prom_file.exists():
    wb=load_workbook(prom_file,data_only=True,read_only=False)
    ws=wb[wb.sheetnames[0]]
    headers=[c.value for c in ws[3]]
    for row in ws.iter_rows(min_row=4,values_only=True):
        if not row or not row[0]: continue
        year=str(row[0]).replace('–','-')
        for idx,h in enumerate(headers[1:],start=1):
            val=row[idx] if idx<len(row) else None
            if isinstance(val,(int,float)):
                promotion.append({'year':year,'form':str(h),'line':float(val),'note':''})
            elif val:
                promotion.append({'year':year,'form':str(h),'line':None,'note':str(val)})
# trend summary
by_form=defaultdict(list)
for r in promotion:
    if r['line'] is not None: by_form[r['form']].append(r)
promo_summary=[]
for form,arr in by_form.items():
    arr=sorted(arr,key=lambda x:x['year'])
    vals=[x['line'] for x in arr]
    promo_summary.append({'form':form,'years':f"{arr[0]['year']}至{arr[-1]['year']}",'n_years':len(vals),'min_line':min(vals),'max_line':max(vals),'avg_line':round(sum(vals)/len(vals),2),'latest_year':arr[-1]['year'],'latest_line':arr[-1]['line'],'change_first_latest':round(arr[-1]['line']-arr[0]['line'],2)})
# Count below latest line using internal exam raw averages (2025/26)
def num(x):
    if x is None: return None
    if isinstance(x,(int,float)) and not isinstance(x,bool): return float(x) if not math.isnan(x) else None
    m=re.match(r'^(\d+(?:\.\d+)?)',str(x).strip())
    return float(m.group(1)) if m else None
below=[]
latest_lines={r['form']:r['line'] for r in promotion if r['year']=='2025-2026' and r['line'] is not None}
base=Path('/home/user/drive_data/Internal Exam/2025_26')
for f in sorted(base.glob('S[1-5].xlsx')):
    form=f.stem
    line=latest_lines.get(form)
    if line is None: continue
    wb=load_workbook(f,data_only=True,read_only=False)
    for sheet in ['Exam1','FT2','Exam2']:
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]; hr=None
        for rr in range(1,min(15,ws.max_row)+1):
            vals=[ws.cell(rr,c).value for c in range(1,ws.max_column+1)]
            if '班別' in vals and '學生姓名' in vals:
                hr=rr; break
        if not hr: continue
        total=0; low=0; avgs=[]
        for row in ws.iter_rows(min_row=hr+1,values_only=True):
            if not row or row[0] is None or row[2] is None: continue
            avg=num(row[5] if len(row)>5 else None)
            if avg is None: continue
            total+=1; avgs.append(avg)
            if avg<line: low+=1
        if total:
            below.append({'year':'2025-2026','form':form,'assessment':sheet,'line':line,'students':total,'below_line':low,'below_pct':pct(low,total),'average_score':round(sum(avgs)/len(avgs),2)})
D['promotion_line']={'records':promotion,'summary':promo_summary,'below_line_2025_26':below,'definition':'升留班線為各級各學年用作檢視學生升級、留級或需進一步支援的校內參考分數線。熱圖顯示不同年度及級別的分數線高低；另以2025/26校內試平均分初步估算低於該級分數線的學生比例。'}
open('/home/user/analysis_output/analysis_data.json','w',encoding='utf-8').write(json.dumps(D,ensure_ascii=False,indent=2))
print('updated degree and promotion', len(promotion), len(below))
