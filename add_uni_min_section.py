import json,re,math,hashlib,statistics
from pathlib import Path
from collections import Counter,defaultdict
from openpyxl import load_workbook
BASE=Path('/home/user/drive_data')
OUT=Path('/home/user/analysis_output/analysis_data.json')
D=json.load(open(OUT,encoding='utf-8'))

def norm(x): return '' if x is None else re.sub(r'\s+','',str(x).strip().upper())
def disp(x): return '' if x is None else str(x).strip()
def num(x):
    if x is None: return None
    if isinstance(x,(int,float)) and not isinstance(x,bool): return float(x) if not math.isnan(x) else None
    if isinstance(x,str):
        s=x.strip().upper()
        if s=='U': return 0.0
        if s in ['A','ATTAINED']: return 'A'
        m=re.match(r'^(\d+(?:\.\d+)?)',s)
        return float(m.group(1)) if m else None
    return None

def pct(a,b): return round(100*a/b,1) if b else None

def subject_columns(ws):
    header_row=None
    for r in range(1,min(6,ws.max_row)+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if '班別' in vals:
            header_row=r; break
    if not header_row: return None,[]
    hdr=[ws.cell(header_row,c).value for c in range(1,ws.max_column+1)]
    grp=[ws.cell(header_row-1,c).value if header_row>1 else None for c in range(1,ws.max_column+1)]
    cols=[]
    for i,h in enumerate(hdr):
        hstr=str(h).strip() if h is not None else ''; gstr=str(grp[i]).strip() if grp[i] is not None else ''
        subj=None
        if gstr.upper()=='CHINESE' and hstr in ['整體','Overall']: subj='中文'
        elif gstr.upper()=='ENGLISH' and hstr in ['Overall','整體','TOTAL']: subj='英文'
        elif hstr in ['數學','公民','M1','M2','會計','商管','生物','化學','物理','中史','歷史','地理','旅款','經濟','ICT','BAFS','企會財','企管','中國文學','視藝']:
            subj=hstr
        if subj: cols.append((i,subj))
    return header_row,cols
# DSE students 2022-2025
dse=[]
for year in [2022,2023,2024,2025]:
    p=BASE/f'HKDSE/各班成績/各班成績{year}.xlsx'
    if not p.exists(): p=BASE/f'HKDSE/Reference/{year}/各班成績{year}.xlsx'
    if not p.exists(): continue
    wb=load_workbook(p,data_only=True,read_only=False); ws=wb[wb.sheetnames[0]]
    hr,cols=subject_columns(ws)
    if not hr: continue
    for row in ws.iter_rows(min_row=hr+1, values_only=True):
        if not row or not row[0] or not row[3]: continue
        scores={}
        for idx,subj in cols:
            v=num(row[idx] if idx<len(row) else None)
            if v is not None: scores[subj]=v
        numeric=[v for v in scores.values() if isinstance(v,(int,float))]
        best5=sum(sorted(numeric,reverse=True)[:5]) if len(numeric)>=5 else None
        chi=scores.get('中文'); eng=scores.get('英文'); mathv=scores.get('數學')
        other2=sum(1 for k,v in scores.items() if k not in ['中文','英文','數學','公民'] and isinstance(v,(int,float)) and v>=2)
        citizenship_ok=(scores.get('公民')=='A' or scores.get('公民')==1 or scores.get('公民') is None) # older years may not have 公民
        min_uni=isinstance(chi,(int,float)) and chi>=3 and isinstance(eng,(int,float)) and eng>=3 and isinstance(mathv,(int,float)) and mathv>=2 and other2>=1 and citizenship_ok
        dse.append({'year':year,'class':disp(row[0]),'no':int(row[1]) if isinstance(row[1],(int,float)) else row[1], 'ename':disp(row[2]),'cname':disp(row[3]), 'key':(year,norm(row[3])), 'scores':scores,'best5':best5,'min_uni':min_uni, 'chi':chi,'eng':eng,'math':mathv,'elective2_count':other2})
# Career by year/name
career_by_key={}
for r in D.get('career',{}).get('linked_records',[]):
    career_by_key[(r['grad_year'], r.get('cname',''))]=r
# linked_records uses normalized cname; build also from records
for r in D.get('career',{}).get('records',[]):
    career_by_key.setdefault((r['grad_year'], r.get('cname','')), r)
records=[]
for s in dse:
    if not s['min_uni']: continue
    c=career_by_key.get((s['year'], norm(s['cname'])), {})
    scores=s['scores']
    records.append({'year':s['year'],'class':s['class'],'no':s['no'],'中文姓名':s['cname'],'英文姓名':s['ename'],'best5':s['best5'],'中文':s['chi'],'英文':s['eng'],'數學':s['math'],'選修2級或以上數目':s['elective2_count'],'出路狀態':c.get('status',''),'課程層級':c.get('level',''),'院校途徑':c.get('institution_type',''),'院校':c.get('institution',''),'課程':c.get('programme',''),'範疇':c.get('field','')})
# Summary by year/outcome/best5 bands
summary=[]
for y in sorted(set(r['year'] for r in records)):
    arr=[r for r in records if r['year']==y]
    total=len(arr)
    summary.append({'year':y,'metric':'符合最低大學入學資格人數','category':'全部','count':total,'pct':100 if total else 0})
    for dim in ['課程層級','院校途徑','範疇']:
        for k,v in Counter(r[dim] or '未分類' for r in arr).most_common():
            summary.append({'year':y,'metric':dim,'category':k,'count':v,'pct':pct(v,total)})
    bands={'10-14':lambda x:x is not None and 10<=x<=14,'15-19':lambda x:x is not None and 15<=x<=19,'20+':lambda x:x is not None and x>=20}
    for b,fn in bands.items():
        n=sum(fn(r['best5']) for r in arr)
        summary.append({'year':y,'metric':'Best5分段','category':b,'count':n,'pct':pct(n,total)})
# Cohort aggregate trend using TSA DSE_year = TSA_year+3 where available + DSE min uni count
trend=[]
for y in sorted(set(r['year'] for r in records)):
    # DSE aggregate min uni count and avg best5
    arr=[r for r in records if r['year']==y]
    trend.append({'dse_year':y,'stage':'DSE','subject':'最低大學入學資格','value':len(arr),'unit':'人數'})
    if arr:
        trend.append({'dse_year':y,'stage':'DSE','subject':'符合資格學生Best5平均','value':round(sum(r['best5'] for r in arr if r['best5'] is not None)/sum(1 for r in arr if r['best5'] is not None),2),'unit':'分'})
    tsa_year=str(y-3)
    for t in D.get('tsa',[]):
        if t.get('year')==tsa_year and 'SEN' not in t.get('subject','') and t.get('school_basic_pct') is not None:
            trend.append({'dse_year':y,'stage':'初中TSA','subject':t['subject'],'value':t['school_basic_pct'],'unit':'基本水平%'})
# Case selection with real names: top best5, diverse outcomes, edge qualifying
cases=[]
for y in sorted(set(r['year'] for r in records)):
    arr=[r for r in records if r['year']==y]
    # top 5 each year
    for r in sorted(arr,key=lambda x:(x['best5'] or 0),reverse=True)[:5]:
        rr=dict(r); rr['個案類型']='高分並符合最低大學入學資格'; rr['解讀']='可檢視其科目組合、升學選擇及對學弟妹的示範價值。'; cases.append(rr)
    # edge: best5 low-ish but qualified
    for r in sorted([x for x in arr if x['best5'] is not None], key=lambda x:x['best5'])[:5]:
        rr=dict(r); rr['個案類型']='剛達資格／邊緣成功'; rr['解讀']='可分析核心科達標策略、選修科組合及升學輔導如何協助其取得可行出路。'; cases.append(rr)
    # non bachelor with qualification
    for r in [x for x in arr if x['課程層級'] not in ['學士/本科','']][:5]:
        rr=dict(r); rr['個案類型']='符合資格但選擇非本科路徑'; rr['解讀']='可了解學生志向、課程興趣及副學士／高級文憑等銜接策略。'; cases.append(rr)
# dedup by year/name/type? keep
D['university_minimum']={'definition':'本部分以DSE成績中「中文3級或以上、英文3級或以上、數學2級或以上，並至少一個選修科2級或以上」作為最低大學入學資格的可量化分析口徑；如該年度有公民科資料，亦按表內達標記錄處理。','records':records,'summary':summary,'trend':trend,'cases':cases[:120]}
open(OUT,'w',encoding='utf-8').write(json.dumps(D,ensure_ascii=False,indent=2))
print('min uni records',len(records),'cases',len(cases),Counter(r['year'] for r in records))
